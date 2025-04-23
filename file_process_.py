# conda activate comic
import os
from PIL import Image, ImageDraw
import matplotlib.pyplot as plt
import requests
import glob
from io import BytesIO
import openpyxl
from tqdm import tqdm
# import demoji
import re
from PIL import Image
from transformers import Blip2Processor, Blip2ForConditionalGeneration
import torch
import json
import cv2
import csv
import onnxruntime as rt
import numpy as np
import shutil


def prepare_images():
    print("1")
    if not os.path.exists(image_save_dir):
        os.makedirs(image_save_dir, exist_ok=True)
    print("2")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Sheet1"]

    # for r in tqdm(range(1,1918)):
    #     image_name = str(ws.cell(row=r,column=image_name_index).value)
    #     image_url = ws.cell(row=r,column=image_url_index).value

    #     loaded_image = load_image(image_url)
    #     loaded_image.save(f"{image_save_dir}/{image_name}.{image_save_postfix}")

    #     # if r == 2:
    #     #    break
    r = 2
    while True:
        image_name = str(ws.cell(row=r, column=image_name_index).value)
        image_url = ws.cell(row=r, column=image_url_index).value

        if not image_name or not image_url:
            break

        image_save_path = f"{image_save_dir}/{image_name}.{image_save_postfix}"
        if os.path.exists(image_save_path):
            continue
        
        loaded_image = load_image(image_url)
        loaded_image.save(image_save_path)
        r += 1

    image_names = os.listdir(image_save_dir)
    print('len(image_names): ', len(image_names))
    # image_names = [image_name for image_name in image_names if image_name not in [".ipynb_checkpoints", ".DS_Store"] and f'.{image_save_postfix}' in image_name]
    
    for image_name in image_names:
        if (image_name in [".ipynb_checkpoints", ".DS_Store"] or f'.{image_save_postfix}' not in image_name):
            try:
                image_path = os.path.join(image_save_dir, image_name)
                os.remove(image_path)
            except OSError as e:
                print(f"Error deleting {image_name}: {e}")

def generate_blip2_captions_json():
    image_names = sorted(os.listdir(image_save_dir))
    with open(MGC_blip2_captions_txt_path, 'w') as file:
        for item in image_names:
            file.write(item + '\n')

# def remove_emojis(text):
#     cleaned_text = demoji.replace(text, '')
#     return cleaned_text

def keep_alphanumeric(text):
    alphanumeric_pattern = re.compile(r'[\W_]+', re.UNICODE)
    text_alphanumeric = alphanumeric_pattern.sub(r'', text)
    return text_alphanumeric

# def load_image(url):
#     try:
#         response = requests.get(url)
#     except:
#         return None
#     return Image.open(BytesIO(response.content)).convert("RGB")

def load_image(url):
    img = None
    timeout_num = 0
    timeout_max_try_num = 3
    load_img_stuck_wait_time = 5
    while True:
        try:
            response = requests.get(url, timeout=30)
            img = Image.open(BytesIO(response.content)).convert("RGB")
            break
        except Exception as e:
            print(f"load image {url} stuck, try to load again after {load_img_stuck_wait_time} seconds", str(e))
            time.sleep(load_img_stuck_wait_time)
            timeout_num += 1
            if timeout_num == timeout_max_try_num:
                print(f"load image {url} stuck, the max try num {timeout_max_try_num} has been reached, check the network and try again", "")
                break
    return img

def generate_blip2_captions_danbooru_():
    device = "cuda" if torch.cuda.is_available() else "cpu"

    processor = Blip2Processor.from_pretrained(Salesforce_blip2_opt_67b_coco_model_path)
    model = Blip2ForConditionalGeneration.from_pretrained(Salesforce_blip2_opt_67b_coco_model_path, torch_dtype=torch.float16)
    model.to(device)

    # image_save_dir = "/nas40/chenyu.liu/fastcomposer_release_danbooru/fastcomposer-main/data/train_fastcomposer_data_336k_pre_release_danbooru_/00000"
    # image_save_dir = "/nas40/chenyu.liu/Datasets/Genshin/Albedo/orig"
    # image_save_dir = "/dfs/comicai/chenyu.liu/Datasets/MGC/00000/"
    captions = {}
    # image_names = [image_name for image_name in image_names if image_name not in [".ipynb_checkpoints", ".DS_Store"] and '.png' in x] # 删除所有特定元素
    image_names = sorted(os.listdir(image_save_dir))
    for image_name in tqdm(image_names):
        file_name = image_name.split('.')[0]
        image_name_path = os.path.join(image_save_dir, image_name)
        raw_image = Image.open(image_name_path)
        inputs = processor(images=raw_image, return_tensors="pt").to(device, torch.float16)
        generated_ids = model.generate(**inputs)
        generated_text = processor.batch_decode(generated_ids, skip_special_tokens=True)[0].strip()
        captions[file_name]= [generated_text]
        # print(image_name, generated_text)
        
    json_str = json.dumps(captions, indent=4)
    with open(blip2_captions_MGC_1916_batch_json_path, 'w') as f:
        lines = json_str.split('\n')
        for line in lines:
            f.write(line)
            f.write('\n')

def generate_image_json():
    image_list = []
    num_no_a = 0
    num_no_gender =0
    with open(blip2_captions_MGC_1916_batch_json_path, 'r') as f:
        data = json.load(f)
        for key, value in data.items():
            # get image_name
            image_name = key
            
            # get caption
            caption = value[0]

            # get word
            caption_list = caption.split()
            sub_str = None
            gender_list = ['girl', 'boy', 'woman', 'man', 'female', 'male', 'character', 'child', 'person', 'bride', 'couple', 'guy', 'robot', 'chef', 'maid', 'witch', 'mermaid', 'princess', 'fairy', 'angel', 'cat', 'dog', 'bird', 'rabbit', 'mouse', 'panda', 'zombie']
            for gender in gender_list:
                if gender in caption_list:
                    start_index = caption_list.index(gender)
                    end_index = None
                    for i in range(start_index-1, -1, -1):
                        if caption_list[i] in ['a', 'an']:
                            end_index = i
                            break
                    if end_index != None:
                        sub_str = " ".join(caption_list[end_index:start_index+1])
                        break
                    else:
                        sub_str = gender
                        # print(f"{image_name}/{caption}中的{gender}前面没有'a'或'an', sub_str就用{gender}代替")
                        num_no_a += 1
                        break

            if sub_str != None:
                word = sub_str
            else:
                # print(f"gender_list中的任意一个gender都不在{image_name}/{caption}中，那就人为添加一个'a girl '在caption前面")
                num_no_gender += 1
                caption = "a girl " + caption
                word = "a girl"

            # get start, end
            start = caption.find(word)
            if start != -1:
                end = start + len(word)
            else:
                print('absolutely impossible:')

            image_list.append([image_name, caption, word, start, end])

    sorted_image_list = sorted(image_list[0:], key=lambda x: x[0])
    print('len(sorted_image_list):', len(sorted_image_list))
    print('num_no_a:', num_no_a)
    print('num_no_gender:', num_no_gender)

    for role_pic in tqdm(sorted_image_list): #[:2]
        role_info = {}
        role_info["image_id"] = role_pic[0]
        role_info["caption"] = role_pic[1] # ' '.join(['a', role_pic[1]])
        ##
        word = role_pic[2]
        start = role_pic[3]
        end = role_pic[4]

        with Image.open(f"{image_save_dir}/{role_info['image_id']}.png") as img:
            width, height = img.size

        role_info["segments"] = [{"id": 1, "word": word, "start": start, "end": end, "bbox": [0, 0, width, height], "coco_label": "person"}]

        # print('role_info: ', role_info)

        with open(f"{image_save_dir}/{role_info['image_id']}_blip2_captions.json", "w") as f:
            json.dump(role_info, f)
            f.write('\n')

def get_mask(img, s=1024):
    img = (img / 255).astype(np.float32)
    h, w = h0, w0 = img.shape[:-1]
    h, w = (s, int(s * w / h)) if h > w else (int(s * h / w), s)
    ph, pw = s - h, s - w
    img_input = np.zeros([s, s, 3], dtype=np.float32)
    img_input[ph // 2:ph // 2 + h, pw // 2:pw // 2 + w] = cv2.resize(img, (w, h))
    img_input = np.transpose(img_input, (2, 0, 1))
    img_input = img_input[np.newaxis, :]
    mask = rmbg_model.run(None, {'img': img_input})[0][0]
    mask = np.transpose(mask, (1, 2, 0))
    mask = mask[ph // 2:ph // 2 + h, pw // 2:pw // 2 + w]
    mask = cv2.resize(mask, (w0, h0))[:, :, np.newaxis]
    return mask

def rmbg_fn(img):
    mask = get_mask(img)
    # img = (mask * img + 255 * (1 - mask)).astype(np.uint8)
    img = (mask * img).astype(np.uint8)
    mask = (mask * 255).astype(np.uint8)
    img = np.concatenate([img, mask], axis=2, dtype=np.uint8)
    mask = mask.repeat(3, axis=2)
    return mask, img

def generate_mask_image():
    if not os.path.exists(target_data_336k_pre_mask_root_dir):
        os.makedirs(target_data_336k_pre_mask_root_dir, exist_ok=True)

    providers = ['CUDAExecutionProvider', 'CPUExecutionProvider']
    ##
    # skytnt_anime_seg_isnetis_onnx_model_path = "extract_mask/isnetis.onnx"
    # skytnt_anime_seg_isnetis_onnx_model_path = "model_ckpt/isnetis.onnx"
    # skytnt_anime_seg_isnetis_onnx_model_path = "/dfs/comicai/chenyu.liu/Models/isnetis/isnetis.onnx"
    rmbg_model = rt.InferenceSession(skytnt_anime_seg_isnetis_onnx_model_path, providers=providers)
    
    # MGC_blip2_captions_txt_path = '/dfs/comicai/chenyu.liu/Datasets/MGC/MGC_blip2_captions.txt'
    role_names = []
    with open(MGC_blip2_captions_txt_path, 'r') as file:
        for line in file:
            role_names.append(line.strip())
    print(role_names)

    # role_names = os.listdir(image_save_dir)
    for i, role_pic in tqdm(enumerate(role_names)):
        # print(f"{i}/{len(role_names)}")
        # id = 1 #表示这张图片中的第id个人
        input_img_path = os.path.join(image_save_dir, role_pic)
        # print("input_img_path: ", input_img_path)
        input_img = cv2.imread(input_img_path)
        output_mask, output_img = rmbg_fn(input_img)
        role_pic_name = role_pic.split('.')[0]
        
        seg_mask_dir = os.path.join(target_data_336k_pre_mask_root_dir, role_pic_name, role_pic_name+'_seg_mask', role_pic_name)
        if not os.path.exists(seg_mask_dir):
            os.makedirs(seg_mask_dir, exist_ok=True)
        cv2.imwrite(os.path.join(seg_mask_dir, role_pic_name+'.jpg'), output_mask)
        # shutil.copy(os.path.join(orig_data_336k_pre_root_dir, role_pic_name+'_mask.jpg'), os.path.join(seg_mask_dir, role_pic_name+'.png'))
        
        sub_seg_dir = os.path.join(target_data_336k_pre_mask_root_dir, role_pic_name, role_pic_name+'_sub_seg', role_pic_name)
        if not os.path.exists(sub_seg_dir):
            os.makedirs(sub_seg_dir, exist_ok=True)
        cv2.imwrite(os.path.join(sub_seg_dir, role_pic_name+'.jpg'), output_img)
        # shutil.copy(os.path.join(orig_data_336k_pre_root_dir, role_pic_name+'_img_with_mask.jpg'), os.path.join(sub_seg_dir, role_pic_name+'.png'))

        # if i == 2:
        #     break
        
def mask2npy(mask_dir, npy_dir):
    mask_image = Image.open(mask_dir).convert("L")
    mask_array = np.array(mask_image)
    mask_array = np.where(mask_array > 130, 1, 0).astype(np.uint8)
    np.save(npy_dir, mask_array)

def npy2mask(npy_dir, mask_dir):
    data = np.load(npy_dir)
    data = np.where(data == 1, 255, 0).astype(np.uint8)
    image = Image.fromarray(data, mode='L')
    image.save(mask_dir)

def generate_npy():
    role_names = os.listdir(image_save_dir)
    for i, role_pic in tqdm(enumerate(role_names)):
        role_pic_name = role_pic.split('.')[0]
        seg_mask_dir = os.path.join(target_data_336k_pre_mask_root_dir, role_pic_name, role_pic_name+'_seg_mask', role_pic_name, role_pic_name+'.jpg')
        target_dir = os.path.join(target_data_336k_pre_mask_root_dir, role_pic_name + '.npy')
        mask2npy(seg_mask_dir, target_dir)

def generate_image_ids_txt():
    IDS = []
    role_names = os.listdir(image_save_dir)
    for i, role_name in tqdm(enumerate(role_names)):
        if 'mask' not in role_name:
            image_name = role_name.split('.')[0]
            IDS.append(image_name)

    IDS.sort()
    id_str = '\n'.join(IDS)

    with open(os.path.join(image_root_dir, "image_ids_train.txt"), "w") as file:
        file.write(id_str)


if __name__ == "__main__":
    image_root_dir = "/data/code/chenyu.liu/Datasets/MGC"
    xlsx_path = os.path.join(image_root_dir, "MGC.xlsx")
    image_name_index = 1
    image_url_index = 2
    image_save_dir = os.path.join(image_root_dir, "00000")
    image_save_postfix = "jpg" # "webp"
    MGC_blip2_captions_txt_path = os.path.join(image_root_dir, "MGC_blip2_captions.txt")
    blip2_captions_MGC_1916_batch_json_path = os.path.join(image_root_dir, "blip2_captions_MGC_1916_batch_.json")
    target_data_336k_pre_mask_root_dir = os.path.join(image_root_dir, "mask_")

    Salesforce_blip2_opt_67b_coco_model_path = "/data/code/chenyu.liu/Models/Salesforce_blip2-opt-6.7b-coco"
    skytnt_anime_seg_isnetis_onnx_model_path = "/data/code/chenyu.liu/Models/skytnt_anime-seg/isnetis.onnx"

    # 1. 准备数据
    prepare_images()

    # 2. 生成xxxxxxx_blip2_captions.json文件
    # 2.0. 先读取root_path下后缀名为".jpg"的image_names，生成MGC_blip2_captions.txt，防止每次都要从文件夹中重新读取的慢性和不稳定性
    generate_blip2_captions_json()

    # 2.1. 然后运行/dfs/comicai/chenyu.liu/fastcomposer_danbooru/generate_blip2_captions_danbooru_.py，生成/dfs/comicai/chenyu.liu/Datasets/MGC/blip2_captions_MGC_1916_batch_.json，TODO: 然后手动将json文件中的' }{'替换为','
    generate_blip2_captions_danbooru_()

    # 2.2. 根据blip2_captions_MGC_1916_batch_.json文件，生成每个图片的json文件到数据集的文件夹下
    generate_image_json()

    # 3. 新建mask_文件夹, 生成mask相关图片
    generate_mask_image()

    # 4. 生成xxxxxxx.npy文件
    # 把所有形如xxxxxxx_mask.jpg的文件，经过mask2npy处理之后，保存为xxxxxxx.npy文件到新建的train_fastcomposer_data_336k_pre_release_danbooru_/00000文件夹下
    generate_npy()

    # 5. 生成image_ids.txt等文件
    # 生成image_ids.txt（暂时不用生成）、image_ids_train.txt、image_ids_test.txt（暂时不用生成）文件
    generate_image_ids_txt()